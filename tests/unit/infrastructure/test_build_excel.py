from __future__ import annotations
from io import BytesIO
from openpyxl import load_workbook
from app.domain.helpdesk import HelpdeskRequest
from app.infrastructure.build_excel import build_excel


# build a HelpdeskRequest
def _make_request(
    id: str = "1",
    request_category: str = "Category A",
    request_type: str = "Type X",
    short_description: str = "Some short description",
    sla_value: int | None = 3,
    sla_unit: str | None = "days",
) -> HelpdeskRequest:
    return HelpdeskRequest(
        id=id,
        request_category=request_category,
        request_type=request_type,
        short_description=short_description,
        sla_value=sla_value,
        sla_unit=sla_unit,
    )

def test_build_excel() -> None:
    # given
    # no LLM or classifier involved
    requests = [
        _make_request(id="1", short_description="first"),
        _make_request(id="2", short_description="second"),
    ]

    # when
    excel_bytes = build_excel(requests)

    # then
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active

    assert ws.title == "Helpdesk Requests"

    # header checks
    header_cell = ws["A1"]
    assert header_cell.value == "id"
    assert header_cell.font.bold is True
    assert header_cell.font.size == 14

    # data cell checks
    data_cell = ws["A2"]
    assert data_cell.value == "1"
    assert data_cell.font.bold is False
    assert data_cell.font.size == 14

    # all cells should have a border set
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=ws.max_column,
    ):
        for cell in row:
            assert cell.border is not None
            # at least one side should have a style
            assert any(
                side.style is not None
                for side in (
                    cell.border.left,
                    cell.border.right,
                    cell.border.top,
                    cell.border.bottom,
                )
            )

# verify hierarchical sorting by category, type, and short_description
def test_build_excel_sorts() -> None:
    # given
    # intentionally unsorted input
    requests = [
        _make_request(
            id="1",
            request_category="Category B",
            request_type="Type Z",
            short_description="zzz",
        ),
        _make_request(
            id="2",
            request_category="Category A",
            request_type="Type Y",
            short_description="beta",
        ),
        _make_request(
            id="3",
            request_category="Category A",
            request_type="Type X",
            short_description="alpha",
        ),
        _make_request(
            id="4",
            request_category="Category A",
            request_type="Type X",
            short_description="aaa",
        ),
    ]

    # when
    excel_bytes = build_excel(requests)

    # then
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active

    # rows start from 2 because row 1 is header
    ids_order = [ws[f"A{row}"].value for row in range(2, 2 + len(requests))]

    # expected order by:
    #   request_category ASC
    #   request_type ASC
    #   short_description ASC
    # → Category A / Type X / "aaa" (4)
    #   Category A / Type X / "alpha" (3)
    #   Category A / Type Y / "beta" (2)
    #   Category B / Type Z / "zzz" (1)
    assert ids_order == ["4", "3", "2", "1"]


# verify columns are auto-fitted based on content length
def test_build_excel_auto_fits_columns() -> None:
    # given
    # second row has much longer short_description
    requests = [
        _make_request(
            id="1",
            short_description="short",
        ),
        _make_request(
            id="2",
            short_description="this is a much longer short description for testing",
        ),
    ]

    # when
    excel_bytes = build_excel(requests)

    # then
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active

    # column A: id, small content
    width_id = ws.column_dimensions["A"].width
    # column D: short_description, larger content
    width_short_desc = ws.column_dimensions["D"].width

    # widths should be numbers and short_description column should be wider
    assert isinstance(width_id, (int, float))
    assert isinstance(width_short_desc, (int, float))
    assert width_short_desc > width_id