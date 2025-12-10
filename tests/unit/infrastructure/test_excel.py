from __future__ import annotations
from io import BytesIO
from openpyxl import load_workbook
from app.domain.helpdesk import HelpdeskRequest
from app.infrastructure.excel import build_excel


# build a HelpdeskRequest
def _make_request(
    raw_id: str = "1",
    request_category: str = "Category A",
    request_type: str = "Type X",
    short_description: str = "Some short description",
    sla_value: int | None = 3,
    sla_unit: str | None = "days",
) -> HelpdeskRequest:
    return HelpdeskRequest(
        raw_id=raw_id,
        request_category=request_category,
        request_type=request_type,
        short_description=short_description,
        sla_value=sla_value,
        sla_unit=sla_unit,
        raw_payload={},                                                                                                     # type: ignore[arg-type]
    )

def test_build_excel() -> None:
    # given
    # no LLM or classifier involved
    requests = [
        _make_request(raw_id="1", short_description="first"),
        _make_request(raw_id="2", short_description="second"),
    ]

    # when
    excel_bytes = build_excel(requests)

    # then
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active

    assert ws.title == "Helpdesk Requests"

    # header checks
    header_cell = ws["A1"]
    assert header_cell.value == "raw_id"
    assert header_cell.font.bold is True
    assert header_cell.font.size == 14

    # data cell checks
    data_cell = ws["A2"]
    assert data_cell.value == "1"
    assert data_cell.font.bold is False
    assert data_cell.font.size == 14

    # all cells should have a border set
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=ws.max_column,
    ):
        for cell in row:
            assert cell.border is not None
            # at least one side should have a style
            assert any(
                side.style is not None
                for side in (
                    cell.border.left,
                    cell.border.right,
                    cell.border.top,
                    cell.border.bottom,
                )
            )